import logging
from openpyxl import load_workbook
import attr
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta
from jinja2 import Environment, FileSystemLoader
from datetime import datetime
from cpd_rules import CPD_TYPES, CPD_MINS

env = Environment(loader=FileSystemLoader('templates'))
template = env.get_template('cpd_summary.html')


def parse_date(date_str):
    """ Parse a date string """
    return parse(date_str, dayfirst=True)


@attr.s
class CPDTotal(object):
    cpd_type = attr.ib()
    risk_hrs = attr.ib()
    bus_hrs = attr.ib()
    area_hrs = attr.ib()
    other_hrs = attr.ib()

    @property
    def total_hrs(self):
        total = self.risk_hrs + self.bus_hrs + self.area_hrs + self.other_hrs
        if not self.cpd_limit:
            return total  # No limits
        if self.cpd_limit > total:
            return total
        else:
            return self.cpd_limit

    @property
    def cpd_desc(self):
        return CPD_TYPES[self.cpd_type]['desc']

    @property
    def cpd_long_desc(self):
        return CPD_TYPES[self.cpd_type]['desc_long']

    @property
    def conditions(self):
        return CPD_TYPES[self.cpd_type]['conditions']

    @property
    def cpd_limit(self):
        return CPD_TYPES[self.cpd_type]['limit']


@attr.s
class Record(object):
    ref_no = attr.ib()
    cpd_type = attr.ib()
    start_date = attr.ib(converter=parse_date)
    end_date = attr.ib(converter=parse_date)
    activity = attr.ib()
    topic = attr.ib()
    provider = attr.ib()
    division = attr.ib()
    location = attr.ib()
    total_hrs = attr.ib()
    risk_hrs = attr.ib()
    bus_hrs = attr.ib()
    area_hrs = attr.ib()
    notes = attr.ib()
    learning_outcome = attr.ib()

    @property
    def yr_grp(self):
        """ Return year grouping of CPD """
        if self.start_date.month <= 6:
            return f'{self.start_date.year}-H1'
        return f'{self.start_date.year}-H2'

    def cpd_current(self, years=3):
        """ Return whether record is in last 3 years """
        diff = datetime.now() - self.start_date
        if diff.days < 365.25 * years:
            return True
        return False


def read_workbook(filename: str):

    wb = load_workbook(filename)
    sheet = wb['CPD Records']

    data = []
    for i in range(1, sheet.max_row + 1):
        row = []
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=i, column=j).value
            row.append(cell)
        data.append(row)

    return data


def parse_ea_records(filename):

    data = read_workbook(filename)
    data.pop(0)  # Remove first row
    data.pop(0)  # Remove headings row

    new_rows = list()
    for row in data[0::2]:
        new_rows.append(row)

    for i, row in enumerate(data[1::2]):
        cell = row[3]
        new_rows[i].append(cell)

    records = list()
    for row in new_rows:
        records.append(Record(*row))
    return records


def group_records_by_cpd_type(records, years=3):
    """ Group CPD records by CPD Type """
    type_group = dict()
    for cpd_type in CPD_TYPES.keys():
        type_group[cpd_type] = list()
    for record in records:
        cpd_type = record.cpd_type.replace('Type ', '')
        if record.cpd_current(years):
            type_group[cpd_type].append(record)
    return type_group


def sum_cpd_hours(records):
    risk_hrs = 0
    bus_hrs = 0
    area_hrs = 0
    total_hrs = 0
    for record in records:
        risk_hrs += record.risk_hrs
        bus_hrs += record.bus_hrs
        area_hrs += record.area_hrs
        total_hrs += record.total_hrs
    other_hrs = total_hrs - risk_hrs - bus_hrs - area_hrs
    return risk_hrs, bus_hrs, area_hrs, other_hrs


def get_cpd_totals(records, years=3):
    """ Group CPD by FY and CPD Type """
    totals = list()
    cpd_group = group_records_by_cpd_type(records, years)
    for cpd_type in cpd_group:
        records = cpd_group[cpd_type]
        risk_hrs, bus_hrs, area_hrs, other_hrs = sum_cpd_hours(records)
        t = CPDTotal(cpd_type, risk_hrs, bus_hrs, area_hrs, other_hrs)
        totals.append(t)
    return totals


def build_summary_table(records, years=3):
    """ Build a CPD summary table """
    summary = dict()
    totals = get_cpd_totals(records, years)
    for row in totals:
        summary[row.cpd_type] = {
            'desc': row.cpd_desc,
            'long_desc': row.cpd_long_desc,
            'risk_hrs': row.risk_hrs,
            'bus_hrs': row.bus_hrs,
            'area_hrs': row.area_hrs,
            'other_hrs': row.other_hrs,
            'total_hrs': row.total_hrs,
            'limit': row.cpd_limit,
            'conditions': row.conditions,
        }
    summary['total'] = {
        'desc': 'Total',
        'risk_hrs': sum(row.risk_hrs for row in totals),
        'bus_hrs': sum(row.bus_hrs for row in totals),
        'area_hrs': sum(row.area_hrs for row in totals),
        'other_hrs': sum(row.other_hrs for row in totals),
        'total_hrs': sum(row.total_hrs for row in totals),
        'limit': 150,
        'conditions': CPD_MINS['conditions'],
    }
    return summary


def yearly_hours(records, years=4):
    """ Get hours by year grouping """
    group_data = dict()
    for record in records:
        if record.cpd_current(years):
            yr_grp = record.yr_grp
            if yr_grp not in group_data.keys():
                group_data[yr_grp] = 0
            total_hrs = record.total_hrs
            group_data[yr_grp] += total_hrs
    return group_data


def build_report(cpd_data={}):
    """ Build HTML report """
    output_html = template.render(**cpd_data)
    with open("build/output.html", "w", encoding='utf-8') as fh:
        fh.write(output_html)
    print('Saved file')


if __name__ == '__main__':

    LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s'
    logger = logging.getLogger()  # Root Logger
    logging.basicConfig(level='INFO', format=LOG_FORMAT)

    records = parse_ea_records(filename='record_20190223.xlsx')
    logging.info('Records passed')
    summary_tbl = build_summary_table(records)
    summary_tbl2 = build_summary_table(records, years=2)
    yr_data = yearly_hours(records)

    cpd_data = {
        'summary_tbl': summary_tbl,
        'summary_tbl2': summary_tbl2,
        'mins': CPD_MINS,
        'yr_data': yr_data,
        'records': records,
    }
    build_report(cpd_data)

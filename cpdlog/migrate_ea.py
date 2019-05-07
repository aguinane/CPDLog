import logging
from typing import List
from openpyxl import load_workbook
from dateutil.parser import parse
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from .model import Activities

EA_CATEGORIES = {
    'Type I': 'A',
    'Type II': 'B',
    'Type III': 'C',
    'Type IV': 'D',
    'Type V': 'E',
    'Type VI': 'F',
    'Type VII': 'G',
    'Type VIII': 'H',
}

log = logging.getLogger(__name__)


def parse_ea_export(filename) -> List['ActivityEA']:
    """ Parse EA CPD Export XLSX file """
    wb = load_workbook(filename)
    sheet = wb['CPD Records']

    data = []
    for i in range(1, sheet.max_row + 1):
        row = []
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=i, column=j).value
            row.append(cell)
        data.append(row)

    data.pop(0)  # Remove first row
    data.pop(0)  # Remove headings row

    new_rows = list()
    for row in data[0::2]:
        new_rows.append(row)

    for i, row in enumerate(data[1::2]):
        cell = row[3]
        new_rows[i].append(cell)

    activities = []
    for row in new_rows:
        activity = ActivityEA(*row)
        activities.append(activity)
    return activities


class ActivityEA:
    def __init__(self, *args):
        self.ext_ref = args[0]
        self.cpd_category = EA_CATEGORIES[args[1]]
        self.start_date = parse(args[2], dayfirst=True)
        self.end_date = parse(args[3], dayfirst=True)
        self.activity = args[4]
        self.topic = args[5]
        self.provider = args[6]
        self.division = args[7]
        self.location = args[8]
        self.total_hrs = args[9]
        self.risk_hrs = args[10]
        self.bus_hrs = args[11]
        self.area_hrs = args[12]
        self.notes = args[13]
        self.learning_outcome = args[14]

    def __repr__(self):
        return "<Activity EA {} {} {}>".format(
            self.ext_ref, self.start_date, self.activity
        )


def import_ea_cpd_activities(db_url: str, filename):
    """ Save records from EA CPD Export to specified database """
    engine = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    session = Session()

    for a in parse_ea_export(filename):

        r = session.query(Activities).filter(Activities.ext_ref == a.ext_ref).first()
        if r:
            log.info('Skipped existing entry %s', a.ext_ref)
            continue

        notes = a.learning_outcome
        if a.notes:
            notes += a.notes

        activity = Activities(
            cpd_category=a.cpd_category,
            start_date=a.start_date,
            end_date=a.start_date,
            act_type=a.activity,
            topic=a.topic,
            provider=a.provider,
            location=a.location,
            duration=a.total_hrs,
            notes=notes,
            ext_ref=a.ext_ref,
        )
        session.add(activity)
    session.commit()

